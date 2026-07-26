"""
data.xlsx 表格预处理脚本
========================
1. 末次月经 → YYYY/M/D 格式，空白保留 NaN
2. 检测日期 → YYYY/M/D 格式
3. 检测孕周 → Nw+N 格式（如 13w → 13w+0）
4. 孕妇BMI → 保留 2 位小数
5. 其他数值列 → 自动检测每列最常见小数位数并统一
"""
import pandas as pd
import numpy as np
import re
from collections import Counter

INPUT_FILE = r"data.xlsx"
OUTPUT_FILE = r"data_processed.xlsx"

# ============================================================
# 工具函数
# ============================================================

def parse_lmp_date(val):
    """解析末次月经：datetime → YYYY/M/D 字符串，NaN 保持 NaN"""
    if pd.isna(val):
        return np.nan
    if hasattr(val, "strftime"):
        return f"{val.year}/{val.month}/{val.day}"
    # 兜底：尝试字符串解析
    try:
        dt = pd.to_datetime(val)
        return f"{dt.year}/{dt.month}/{dt.day}"
    except Exception:
        return val  # 无法解析则保留原值


def parse_test_date(val):
    """解析检测日期：YYYYMMDD 整数 → YYYY/M/D 字符串"""
    if pd.isna(val):
        return np.nan
    try:
        s = str(int(val))
        if len(s) == 8 and s.isdigit():
            y, m, d = int(s[:4]), int(s[4:6]), int(s[6:8])
            return f"{y}/{m}/{d}"
    except (ValueError, TypeError):
        pass
    # 兜底
    try:
        dt = pd.to_datetime(val)
        return f"{dt.year}/{dt.month}/{dt.day}"
    except Exception:
        return val


def parse_gestational_week(val):
    """解析孕周 → Nw+N 格式（不补零），如 13w → 13w+0"""
    if pd.isna(val):
        return np.nan

    s = str(val).strip()

    # 已是标准格式: Nw+N
    m = re.match(r"^(\d{1,2})w\+(\d{1,2})$", s, re.IGNORECASE)
    if m:
        return f"{int(m.group(1))}w+{int(m.group(2))}"

    # N周+N
    m = re.match(r"^(\d{1,2})周\+(\d{1,2})$", s)
    if m:
        return f"{int(m.group(1))}w+{int(m.group(2))}"

    # N+N (无单位)
    m = re.match(r"^(\d{1,2})\+(\d{1,2})$", s)
    if m:
        return f"{int(m.group(1))}w+{int(m.group(2))}"

    # NWND / NwNd
    m = re.match(r"^(\d{1,2})[wW](\d{1,2})[dD]$", s)
    if m:
        return f"{int(m.group(1))}w+{int(m.group(2))}"

    # Nw (无+天数，如 13w, 23w)
    m = re.match(r"^(\d{1,2})[wW]$", s)
    if m:
        return f"{int(m.group(1))}w+0"

    # N周 (无+天数)
    m = re.match(r"^(\d{1,2})周$", s)
    if m:
        return f"{int(m.group(1))}w+0"

    # N.D 格式（如 12.3 = 12w+3）
    m = re.match(r"^(\d{1,2})\.(\d{1,2})$", s)
    if m:
        weeks = int(m.group(1))
        days = int(m.group(2))
        if days <= 6 and weeks <= 45:
            return f"{weeks}w+{days}"

    # 纯整数（如 12 = 12w）
    m = re.match(r"^(\d{1,2})$", s)
    if m:
        weeks = int(m.group(1))
        if weeks <= 45:
            return f"{weeks}w+0"

    # 兜底：返回原值
    return s


def detect_decimal_places(series, max_cap=10):
    """检测数值列应保留的小数位数。

    策略：取所有非零小数值的「最大」小数位数（保留最高精度），
    但不超过 max_cap（默认 10，防止浮点噪声如 0.8003690999999999）。
    若所有值都是整数/ .0 结尾，返回 0（后续转为 Int64）。
    """
    s = series.dropna()
    if len(s) == 0:
        return 0

    max_dp = 0
    for v in s:
        if v != int(v):  # 有非零小数部分
            v_str = str(v)
            if "." in v_str:
                d = len(v_str.split(".")[1])
                if max_cap >= d > max_dp:
                    max_dp = d

    return max_dp


# ============================================================
# 主处理流程
# ============================================================

print("=" * 60)
print("开始处理 data.xlsx ...")
print("=" * 60)

xls = pd.ExcelFile(INPUT_FILE)
all_sheets = {}
col_formats = {}  # {sheet_name: {col_name: target_dp}}  用于后续 Excel 格式化

for sheet_name in xls.sheet_names:
    print(f"\n{'─' * 60}")
    print(f"处理 Sheet: [{sheet_name}]")
    df = pd.read_excel(INPUT_FILE, sheet_name=sheet_name)
    print(f"  原始: {df.shape[0]} 行 × {df.shape[1]} 列")

    changes = []
    sheet_fmts = {}  # 当前 sheet 的列格式

    # ----- 标准化列名（去首尾空格）-----
    df.columns = [c.strip() for c in df.columns]

    # ----- 1. 末次月经 → YYYY/M/D -----
    if "末次月经" in df.columns:
        before_nulls = df["末次月经"].isna().sum()
        df["末次月经"] = df["末次月经"].apply(parse_lmp_date)
        after_nulls = df["末次月经"].isna().sum()
        print(f"  ✅ 末次月经: 日期格式化 (NaN: {before_nulls}→{after_nulls})")
        changes.append(f"末次月经 → YYYY/M/D")

    # ----- 2. 检测日期 → YYYY/M/D -----
    if "检测日期" in df.columns:
        df["检测日期"] = df["检测日期"].apply(parse_test_date)
        print(f"  ✅ 检测日期: YYYYMMDD→YYYY/M/D")
        changes.append(f"检测日期 → YYYY/M/D")

    # ----- 3. 检测孕周 → Nw+N -----
    if "检测孕周" in df.columns:
        before_vals = set(df["检测孕周"].dropna().unique())
        df["检测孕周"] = df["检测孕周"].apply(parse_gestational_week)
        after_vals = set(df["检测孕周"].dropna().unique())
        changed = before_vals - after_vals
        if changed:
            print(f"  ✅ 检测孕周: 统一为 Nw+N 格式 (修正 {len(changed)} 种格式)")
            print(f"     示例修正: {list(changed)[:5]}")
        else:
            print(f"  ✅ 检测孕周: 已是 Nw+N 格式，无需修正")
        changes.append(f"检测孕周 → Nw+N")

    # ----- 4. 孕妇BMI → 2 位小数 -----
    if "孕妇BMI" in df.columns:
        df["孕妇BMI"] = pd.to_numeric(df["孕妇BMI"], errors="coerce")
        df["孕妇BMI"] = df["孕妇BMI"].round(2)
        sheet_fmts["孕妇BMI"] = 2  # 记录：需要 2 位小数格式
        print(f"  ✅ 孕妇BMI: 统一为 2 位小数")
        changes.append(f"孕妇BMI → 2 位小数")

    # ----- 5. 其他数值列精度统一 -----
    skip_cols = {"序号", "年龄", "检测抽血次数", "原始读段数", "唯一比对的读段数",
                 "生产次数", "孕妇BMI", "末次月经", "检测日期", "检测孕周"}

    numeric_cols = df.select_dtypes(include=["number"]).columns
    for col in numeric_cols:
        if col in skip_cols:
            continue
        if col.startswith("Unnamed"):
            continue
        if df[col].isna().all():
            continue

        target_dp = detect_decimal_places(df[col])
        if target_dp == 0:
            if (df[col].dropna() % 1 == 0).all():
                df[col] = df[col].astype("Int64")
                changes.append(f"{col} → 整数")
                continue
        df[col] = df[col].round(target_dp)
        sheet_fmts[col] = target_dp  # 记录格式
        changes.append(f"{col} → {target_dp} 位小数")

    print(f"  共计 {len(changes)} 项调整")

    all_sheets[sheet_name] = df
    col_formats[sheet_name] = sheet_fmts

# ============================================================
# 写出结果
# ============================================================
print(f"\n{'=' * 60}")
print(f"写入 {OUTPUT_FILE} ...")

with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
    for sheet_name, df in all_sheets.items():
        df.to_excel(writer, sheet_name=sheet_name, index=False)

    # ---------- 后处理：统一各列 Excel 数字格式 ----------
    from openpyxl.utils import get_column_letter
    for sheet_name, fmts in col_formats.items():
        ws = writer.sheets[sheet_name]
        df = all_sheets[sheet_name]
        for col_name, dp in fmts.items():
            if col_name not in df.columns:
                continue
            col_idx = list(df.columns).index(col_name) + 1  # 1-based
            col_letter = get_column_letter(col_idx)
            fmt_str = "0." + "0" * dp
            for row in range(2, len(df) + 2):  # row 1 = header
                ws[f"{col_letter}{row}"].number_format = fmt_str

print("✅ 完成！")
print(f"\n输出文件: {OUTPUT_FILE}")
print(f"Sheet 数: {len(all_sheets)}")
for sn, df in all_sheets.items():
    print(f"  [{sn}] {df.shape[0]} 行 × {df.shape[1]} 列")
